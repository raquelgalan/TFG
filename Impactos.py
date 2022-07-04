# TFG Raquel Galán Montes
## Cálculo del Aislamiento a Ruido de Impactos UNE-EN ISO 16283-2:2018 ISO 717-2:2013

from openpyxl import load_workbook      # Permite leer ficheros .xlsx
import math                             # Permite usar funciones matemáticas
from pylab import *                     # Permite crear gráficas
from TR import *                        # Se importa el TR del programa

# VARIABLES GLOBALES
APARTADO = 'Datos'                      # Hoja del archivo XLSX
FILE_Impactos = 'Ruido_Impactos.xlsx'   # Variable para el fichero principal
T = 0.5                                 # Tiempo de referencia = 0.5s
V = 23.4                                # Volumen del recinto receptor (V = largo x ancho x alto = 4.5 x 2.08 x 2.5 m^3)

# ARRAY con constantes
## Rango de frecuencias de interes:
arrayFR = [50, 63, 80, 100, 125, 160, 200, 250, 315, 400, 500, 630, 800, 1000, 1250, 1600, 2000, 2500, 3150, 4000, 5000]

# Arrays vacíos
L_Impactos_F1 = []        # Array para los resultados del nivel de ruido de impactos en recepción de la fuente en la posición 1
L_Impactos_F2 = []        # Array para los resultados del nivel de ruido de impactos en recepción de la fuente en la posición 2
L_Impactos_F3 = []        # Array para los resultados del nivel de ruido de impactos en recepción de la fuente en la posición 3
L_Impactos_F4 = []        # Array para los resultados del nivel de ruido de impactos en recepción de la fuente en la posición 4
TR = []                  # Array para los resultados del tiempo de reverberación en recepción
LpRF1 = []               # Array para los resultados del nivel de ruido de fondo en la posición 1
LpRF2 = []               # Array para los resultados del nivel de ruido de fondo en la posición 2
LpRF3 = []               # Array para los resultados del nivel de ruido de fondo en la posición 3
LpRF4 = []               # Array para los resultados del nivel de ruido de fondo en la posición 4
LpRC_F1 = []             # Array para los resultados del nivel de ruido de impactos corregido en la posición 1
LpRC_F2 = []             # Array para los resultados del nivel de ruido de impactos corregido en la posición 2
LpRC_F3 = []             # Array para los resultados del nivel de ruido de impactos corregido en la posición 3
LpRC_F4 = []             # Array para los resultados del nivel de ruido de impactos corregido en la posición 4
LnT_F1 = []              # Array para los resultados del nivel de presión acústica estandarizado en la posición 1
LnT_F2 = []              # Array para los resultados del nivel de presión acústica estandarizado en la posición 2
LnT_F3 = []              # Array para los resultados del nivel de presión acústica estandarizado en la posición 3
LnT_F4 = []              # Array para los resultados del nivel de presión acústica estandarizado en la posición 4
LnT = []                 # Array para los resultados del nivel de presión acústica estandarizado de todas las posiciones

## PROCEDIMIENTOS ESPECÍFICOS DE LA ISO 16283-2
# Cálculo del promedio de los niveles de ruido de impactos con posiciones de micrófono
# en cada banda de frecuencia, Lp, donde A: min_row; B: max_row; C: min_col; D: max_col; n: posiciones de micrófono

def Calcular_L(A, B, C, D, n, myArray):
    wb = load_workbook(FILE_Impactos)        # Se carga en wb el fichero
    apartado = wb[APARTADO]                  # Se carga la hoja del fichero de donde obtenemos los datos

    for value in apartado.iter_rows(min_row =A, max_row =B, min_col = C, max_col =D,
                                 values_only=True):
        i = 0
        L = 0
        for cell in value:
            L = L + 10**(value[i]/10)       # Se calcula de 10*(Lp/10) y se añade al resultado anterior
            i = i + 1
        L = 10*math.log(L/n,(10))           # Se calcula el promedio logarítmico de la suma total
        myArray.append(round(L,1))          # Array ordenado de datos                 

# Corrección del ruido de fondo:
def LCorregido(A, B, myArray):
    i = 0
    h = 10
    l = 6
    correccion = 1.3
    for num in A:
        dif = A[i] - B[i]
        if dif >= h:
            corregido = A[i]
            myArray.append(round(corregido,2))
            i = i + 1
        elif dif <= l:
            corregido_menor = A[i] - correccion
            myArray.append(round(corregido_menor,2))
            i = i + 1
        else:
            corregido_mayor = (10*math.log(10**(A[i]/10)-10**(B[i]/10),10))  
            myArray.append(round(corregido_mayor,1))
            i = i + 1

# Cálculo del nivel de preión acústica estandarizado de ruido de impactos para cada banda de frecuencia, LnT':
# A representa el Li, nivel de presión acustica de ruido de impactos
def Calcular_LnT(A, TR, myArray):
    i = 0
    for num in A:
        LnT = (A[i]) - 10*math.log(TR[i]/T,(10))
        i = i + 1
        myArray.append(round(LnT,1))

# Cálculo del sumatorio de las magnitudes en función de la posición de altavoz:
# donde A, B, C y D son los resultados por cada posición de altavoz respectivamente para cada banda de frecuencias.
def Sumatorio(A, B, C, D, myArray):
    i = 0
    for num in A:
        sum = 10**(A[i]/10) + 10**(B[i]/10) + 10**(C[i]/10) + 10**(D[i]/10)
        X = 10*math.log(sum/4,(10))                # Se divide entre 2, que son el número de posiciones de altavoz
        i = i + 1
        myArray.append(round(X,1))

# Se preparan los resultados para imprimirlos por pantalla
def Valores(myArray, unidades):
    i = 0                                             # Se inicializa el iterador de los arrays
    for num in arrayFR:
        print(arrayFR[i], 'Hz - ', myArray[i], 'dB')  # Se imprime por la terminal el elemento i del array
        i = i + 1                                     # Siguiente elemento

# Datos mostrados en una gráfica
def Representacion_Nivel(A, B):
    title(A)
    xlabel('Frecuencia [Hz]')
    ylabel('Niveles [dB]')
    ylim(0, 110)
    legend((B),
    prop = {'size': 10}, bbox_to_anchor=(1.05, 1.0), loc='best')
    grid()

# Se imprimen todos los datos
if __name__ == "__main__":

    print()
    print('L | MÁQUINA DE IMPACTOS | POSICIÓN 1 DE LA FUENTE')
    print('---------------------------------------------------')
    Calcular_L(33, 53, 7, 8, 2, L_Impactos_F1)
    Valores(L_Impactos_F1, 'dB')

    print()
    print('L | MÁQUINA DE IMPACTOS | POSICIÓN 2 DE LA FUENTE')
    print('---------------------------------------------------')
    Calcular_L(33, 53, 9, 10, 2, L_Impactos_F2)
    Valores(L_Impactos_F2, 'dB')

    print()
    print('L | MÁQUINA DE IMPACTOS | POSICIÓN 3 DE LA FUENTE')
    print('---------------------------------------------------')
    Calcular_L(33, 53, 11, 12, 2, L_Impactos_F3)
    Valores(L_Impactos_F3, 'dB')

    print()
    print('L | MÁQUINA DE IMPACTOS | POSICIÓN 4 DE LA FUENTE')
    print('---------------------------------------------------')
    Calcular_L(33, 53, 13, 14, 2, L_Impactos_F4)
    Valores(L_Impactos_F4, 'dB')

    print()
    print('RUIDO DE FONDO EN HABITACIÓN | POSICIÓN 1 DE LA FUENTE')
    print('-----------------------------------------------------------------')
    Calcular_L(8, 28, 7, 8, 2, LpRF1)
    Valores(LpRF1, 'dB')

    print()
    print('RUIDO DE FONDO EN HABITACIÓN | POSICIÓN 2 DE LA FUENTE')
    print('-----------------------------------------------------------------')
    Calcular_L(8, 28, 10, 11, 2, LpRF2)
    Valores(LpRF2, 'dB')

    print()
    print('RUIDO DE FONDO EN HABITACIÓN | POSICIÓN 3 DE LA FUENTE')
    print('-----------------------------------------------------------------')
    Calcular_L(8, 28, 13, 14, 2, LpRF3)
    Valores(LpRF3, 'dB')

    print()
    print('RUIDO DE FONDO EN HABITACIÓN | POSICIÓN 4 DE LA FUENTE')
    print('-----------------------------------------------------------------')
    Calcular_L(8, 28, 16, 17, 2, LpRF4)
    Valores(LpRF4, 'dB')

    print()
    print('NIVELES CORREGIDOS EN RECEPCIÓN - HABITACIÓN | POSICIÓN 1 DE LA FUENTE')
    print('------------------------------------------------------------------------')
    print('Frecuencia | Lp Corregido Habitación')
    print('-------------------------------------')
    LCorregido(L_Impactos_F1, LpRF1, LpRC_F1)
    Valores(LpRC_F1, 'dB')

    print()
    print('NIVELES CORREGIDOS EN RECEPCIÓN - HABITACIÓN | POSICIÓN 2 DE LA FUENTE')
    print('-------------------------------------------------------------------------')
    print('Frecuencia | Lp Corregido Habitación')
    print('--------------------------------------')
    LCorregido(L_Impactos_F2, LpRF2, LpRC_F2)
    Valores(LpRC_F2, 'dB')

    print()
    print('NIVELES CORREGIDOS EN RECEPCIÓN - HABITACIÓN | POSICIÓN 3 DE LA FUENTE')
    print('-------------------------------------------------------------------------')
    print('Frecuencia | Lp Corregido Habitación')
    print('--------------------------------------')
    LCorregido(L_Impactos_F3, LpRF3, LpRC_F3)
    Valores(LpRC_F3, 'dB')

    print()
    print('NIVELES CORREGIDOS EN RECEPCIÓN - HABITACIÓN | POSICIÓN 4 DE LA FUENTE')
    print('------------------------------------------------------------------------')
    print('Frecuencia | Lp Corregido Habitación')
    print('--------------------------------------')
    LCorregido(L_Impactos_F4, LpRF4, LpRC_F4)
    Valores(LpRC_F4, 'dB')

    print()
    print('Tiempo de Reverberacion de la HABITACIÓN RECEPTORA')
    print('----------------------------------------------------')
    print('Frecuencia | TR')
    print('----------------------------------------------------')
    Calcular_TR(6, 26, 12, 15, TR_F1)
    Calcular_TR(6, 26, 19, 22, TR_F2)
    TR_Habitación(TR_F1, TR_F2, TR)
    Imprimir(TR, 's')

    print()
    print('NIVEL DE PRESIÓN ACÚSTICA ESTANDARIZADO | POSICIÓN 1 DE LA FUENTE')
    print('-----------------------------------------------------------------')
    Calcular_LnT(LpRC_F1, TR, LnT_F1)
    Valores(LnT_F1, 'dB')

    print()
    print('NIVEL DE PRESIÓN ACÚSTICA ESTANDARIZADO | POSICIÓN 2 DE LA FUENTE')
    print('-----------------------------------------------------------------')
    Calcular_LnT(LpRC_F2, TR, LnT_F2)
    Valores(LnT_F2, 'dB')

    print()
    print('NIVEL DE PRESIÓN ACÚSTICA ESTANDARIZADO | POSICIÓN 3 DE LA FUENTE')
    print('-----------------------------------------------------------------')
    Calcular_LnT(LpRC_F3, TR, LnT_F3)
    Valores(LnT_F3, 'dB')

    print()
    print('NIVEL DE PRESIÓN ACÚSTICA ESTANDARIZADO | POSICIÓN 4 DE LA FUENTE')
    print('-----------------------------------------------------------------')
    Calcular_LnT(LpRC_F4, TR, LnT_F4)
    Valores(LnT_F4, 'dB')

    print()
    print('NIVEL DE PRESIÓN ACÚSTICA ESTANDARIZADO PROMEDIADO')
    print('--------------------------')
    Sumatorio(LnT_F1, LnT_F2, LnT_F3, LnT_F4, LnT)
    Valores(LnT, 'dB')

    #REPRESENTACIÓN DE LpRC'
    figure('Niveles de presión acústica de ruido de impactos corregido y estandarizado')
    subplot(3,1,1)
    plot(FR, LpRC_F1, 'co-', LpRC_F2, 'bo-', LpRC_F3, 'm*-', LpRC_F4, 'ks-')            # Genera el gráfico
    Representacion_Nivel('Niveles de presión acústica de ruido de impactos corregido',
    ('LpRC_F1', 'LpRC_F2', 'LpRC_F3', 'LpRC_F4'))

    #REPRESENTACIÓN DE LnT'
    subplot(3,1,3)
    plot(FR, LnT_F1, 'co-', LnT_F2, 'bo-', LnT_F3, 'm*-', LnT_F4, 'ks-', LnT, 'ys-')    # Genera el gráfico
    Representacion_Nivel('Niveles de presión acústica estandarizado',
    ('LnT_F1', 'LnT_F2', 'LnT_F3', 'LnT_F4', 'LnT'))

    tight_layout()                                                                      # Ajusta la leyenda
    show()