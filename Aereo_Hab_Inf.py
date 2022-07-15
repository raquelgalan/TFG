# TFG Raquel Galán Montes
## Cálculo del Aislamiento a Ruido Aéreo  UNE-EN ISO 16283-1:2015 ISO 717-1:2013
### HABITACIÓN INFERIOR, SALA COLINDANTE INFERIOR CON DIVISIÓN HORIZONTAL (SALA PERTENECIENTE A VECINOS)

from openpyxl import load_workbook      # Permite leer de ficheros .xlsx
import math                             # Permite usar funciones matemáticas
from pylab import *                     # Permite crear gráficas
from TR import *                        # Se importa el TR del programa

# VARIABLES GLOBALES
SHEET = 'Datos'                 # Hoja del archivo XLSX
FILE  = 'Ruido_Aereo.xlsx'      # Variable para el fichero principal
T = 0.5                         # Tiempo de referencia = 0.5s
V = 23.4                        # Volumen del recinto receptor (V = largo x ancho x alto = 4.5 x 2.08 x 2.5 m^3)
S = 9.36                        # Superficie del elemento separador (S = suelo = largo x ancho = 4.5 x 2.08 m^2)
C = 0.16                        # Constante para hallar Abs, la superficie de absorción equivalente

# Rango de frecuencias de interes:
arrayFR = [50, 63, 80, 100, 125, 160, 200, 250, 315, 400, 500, 630, 800, 1000, 1250, 1600, 2000, 2500, 3150, 4000, 5000]

# Arrays vacíos
LpR_Inf_F1 = []           # Array para los resultados del nivel de presión en recepción
LpR_Inf_F2 = []           # Array para los resultados del nivel de presión en recepción
LpRC_Inf_F1 = []          # Array para los resultados del nivel de presión en recepción corregido el RF
LpRC_Inf_F2 = []          # Array para los resultados del nivel de presión en recepción corregido el RF
LpRF_Inf = []             # Array para los resultados del nivel de presión del ruido de fondo
LpE_Inf_F1 = []           # Array para los resultados del nivel de presión en emisión de la fuente en la posición 1
LpE_Inf_F2 = []           # Array para los resultados del nivel de presión en emisión de la fuente en posición 2
TR = []                   # Array para los resultados del tiempo de reverberación en recepción
DnT_F1_Inf = []           # Array para los resultados del nivel de diferencia normalizada de la fuente en la posición 1
DnT_F2_Inf = []           # Array para los resultados del nivel de diferencia normalizada de la fuente en la posición 2
DnT_Inf = []              # Array para los resultados del nivel de diferencia normalizada global estandarizada
R_F1_Inf = []             # Array para los resultados del índice de reducción sonora aparente de la fuente en la posición 1
R_F2_Inf = []             # Array para los resultados del índice de reducción sonora aparente de la fuente en la posición 2
R_Inf = []                # Array para los resultados del índice de reducción sonora aparente global, R'

## PROCEDIMIENTOS ESPECÍFICOS DE LA ISO 16283-1
# Cálculo del promedio de posiciones de micrófono para cada banda de frecuencia, Lp, donde
# A: min_row; B: max_row; C: min_col; D: max_col; n: posiciones de micrófono
def Calcular_Lp(A, B, C, D, n, myArray):
    wb = load_workbook(FILE)        # Se carga en wb el fichero
    sheet = wb[SHEET]               # Se carga la hoja del fichero de donde obtenemos los datos

    for value in sheet.iter_rows(min_row = A, max_row = B, min_col = C, max_col = D,
                                 values_only= True):
        i = 0
        L = 0
        for cell in value:
            L = L + 10**(value[i]/10)       # Se calcula de 10*(Lp/10) y se añade al resultado anterior
            i = i + 1
        L = 10*math.log(L/n,(10))           # Se calcula el promedio logarítmico de la suma total
        myArray.append(round(L,1))          # Array ordenado de datos

# Corrección del ruido de fondo:
def LpCorregido(A, B, myArray):
    i = 0
    h = 10
    l = 6
    correccion = 1.3
    for num in A:
        dif = A[i] - B[i]
        if dif >= h:
            corregido = A[i]
            myArray.append(round(corregido,2))
            i = i + 1
        elif dif <= l:
            corregido_menor = A[i] - correccion
            myArray.append(round(corregido_menor,2))
            i = i + 1
        else:
            corregido_mayor = (10*math.log(10**(A[i]/10)-10**(B[i]/10),10))  
            myArray.append(round(corregido_mayor,1))
            i = i + 1

# Cálculo de la diferencia estandarizada para cada banda de frecuencia, DnT:
# A representa el Lp en el recinto EMISOR y B el Lp en el recinto RECEPTOR.
def Diferencia_Nivel(A, B, TR, myArray):
    i = 0
    for num in A:
        diferencia = A[i] - B[i]
        DnT = diferencia + (10*math.log(TR[i]/T,10))  
        i = i + 1
        myArray.append(round(DnT,1))

# Cálculo del sumatorio de las magnitudes en función de la posición de altavoz:
# donde A y B son los resultados por cada posición de altavoz respectivamente para cada banda de frecuencias.
def Sumatorio(A, B, myArray):
    i = 0
    for num in A:
        sum = 10**(-A[i]/10) + 10**(-B[i]/10)
        X = -10*math.log(sum/2,(10))                # Se divide entre 2, que son el número de posiciones de altavoz
        i = i + 1
        myArray.append(round(X,1))


# Se preparan los resultados para imprimirlos por pantalla
def Resultados(myArray, unidades):
    i = 0                                             # Se inicializa el iterador de los arrays
    for num in arrayFR:
        print(arrayFR[i], 'Hz - ', myArray[i], 'dB')  # Se imprime por la terminal el elemento i del array
        i = i + 1                                     # Se va al siguiente elemento

# Cálculo de la reducción sonora aparente para cada banda de frecuencia, R':
# A representa el Lp en el recinto EMISOR y B el Lp en el recinto RECEPTOR.
def Calcular_R(A, B, TR, myArray):
    i = 0
    for num in A:
        Abs = (C * V)/TR[i]                           # Cálculo de la superficie de absorción equivalente
        R = (A[i]-B[i]) + 10*math.log(S/Abs,(10))
        i = i + 1
        myArray.append(round(R,1))

# Datos mostrados en una gráfica
def Representacion_Nivel(A, B):
    title(A)
    xlabel('Frecuencia [Hz]')
    ylabel('Niveles [dB]')
    ylim(0, 110)
    legend((B),
    prop = {'size': 10}, bbox_to_anchor=(1.05, 1.0), loc='best')
    grid()

# Se realizan los cálculos y se imprime por pantalla
if __name__ == "__main__":
    print()
    print('NIVELES EN RECEPCIÓN HABITACIÓN INFERIOR - FUENTE 1')
    print('------------------------------------------------------')
    print('Frecuencia | Lp Hab. Inferior')
    print('------------------------------------------------------')
    Calcular_Lp(113, 133, 4, 8, 5, LpR_Inf_F1)
    Resultados(LpR_Inf_F1, 'dB')

    print()
    print('NIVELES EN RECEPCIÓN HABITACIÓN INFERIOR - FUENTE 2')
    print('-------------------------------------------------------')
    print('Frecuencia | Lp Hab. Inferior')
    print('-------------------------------------------------------')
    Calcular_Lp(113, 133, 9, 13, 5, LpR_Inf_F2)
    Resultados(LpR_Inf_F2, 'dB')
    
    print()
    print('RUIDO DE FONDO EN HABITACIÓN INFERIOR')
    print('----------------------------------------')
    Calcular_Lp(139, 159, 8, 8, 1, LpRF_Inf)
    Resultados(LpRF_Inf, 'dB')

    print()
    print('NIVELES CORREGIDOS EN RECEPCIÓN - HABITACIÓN INFERIOR - FUENTE 1')
    print('------------------------------------------------------------------')
    print('Frecuencia | Lp Corregido Hab. Inferior')
    print('------------------------------------------------------------------')
    LpCorregido(LpR_Inf_F1, LpRF_Inf, LpRC_Inf_F1)
    Resultados(LpRC_Inf_F1, 'dB')

    print()
    print('NIVELES CORREGIDOS EN RECEPCIÓN - HABITACIÓN INFERIOR - FUENTE 2')
    print('-------------------------------------------------------------------')
    print('Frecuencia | Lp Corregido Hab. Inferior')
    print('-------------------------------------------------------------------')
    LpCorregido(LpR_Inf_F2, LpRF_Inf, LpRC_Inf_F2)
    Resultados(LpRC_Inf_F2, 'dB')

    print()
    print('NIVELES EN EMISIÓN HABITACIÓN INFERIOR - FUENTE 1')
    print('---------------------------------------------------')
    print('Frecuencia | Lp Hab. Inferior')
    print('---------------------------------------------------')
    Calcular_Lp(113, 133, 17, 19, 3, LpE_Inf_F1)
    Resultados(LpE_Inf_F1, 'dB')

    print()
    print('NIVELES EN EMISIÓN HABITACIÓN INFERIOR - FUENTE 2')
    print('---------------------------------------------------')
    print('Frecuencia | Lp Hab. Inferior')
    print('---------------------------------------------------')
    Calcular_Lp(113, 133, 20, 22, 3, LpE_Inf_F2)
    Resultados(LpE_Inf_F2, 'dB')

    print()
    print('Tiempo de Reverberacion de la HABITACIÓN RECEPTORA')
    print('----------------------------------------------------')
    print('Frecuencia | TR')
    print('----------------------------------------------------')
    Calcular_TR(6, 26, 12, 15, TR_F1)
    Calcular_TR(6, 26, 19, 22, TR_F2)
    TR_Habitación(TR_F1, TR_F2, TR)
    Imprimir(TR, 's')

    print()
    print('DIFERENCIA DE NIVEL - HABITACIÓN INFERIOR - FUENTE 1')
    print('------------------------------------------------------------------')
    print('Frecuencia | DnT1 Hab. Inferior')
    print('------------------------------------------------------------------')
    Diferencia_Nivel(LpE_Inf_F1, LpRC_Inf_F1, TR, DnT_F1_Inf)
    Resultados(DnT_F1_Inf, 'dB')

    print()
    print('DIFERENCIA DE NIVEL - HABITACIÓN INFERIOR - FUENTE 2')
    print('-------------------------------------------------------------------')
    print('Frecuencia | DnT2 Hab. Inferior')
    print('-------------------------------------------------------------------')
    Diferencia_Nivel(LpE_Inf_F2, LpRC_Inf_F2, TR, DnT_F2_Inf)
    Resultados(DnT_F2_Inf, 'dB')

    print()
    print('DIFERENCIA ESTANDARIZADA')
    print('--------------------------')
    Sumatorio(DnT_F1_Inf, DnT_F2_Inf, DnT_Inf)
    Resultados(DnT_Inf, 'dB')

    print()
    print('ÍNDICE DE REDUCCIÓN SONORA APARENTE FUENTE 1')
    print('----------------------------------------------')
    Calcular_R(LpE_Inf_F1, LpRC_Inf_F1, TR, R_F1_Inf)
    Resultados(R_F1_Inf, 'dB')

    print()
    print('ÍNDICE DE REDUCCIÓN SONORA APARENTE FUENTE 2')
    print('----------------------------------------------')
    Calcular_R(LpE_Inf_F2, LpRC_Inf_F2, TR, R_F2_Inf)
    Resultados(R_F2_Inf, 'dB')

    print()
    print('ÍNDICE DE REDUCCIÓN SONORA APARENTE')
    print('-------------------------------------')
    Sumatorio(R_F1_Inf, R_F2_Inf, R_Inf)
    Resultados(R_Inf, 'dB')

    #REPRESENTACIÓN DE DnT
    figure("DnT de la  Habitación Inferior")
    plot(FR, DnT_F1_Inf, 'co-', DnT_F2_Inf, 'bo-', DnT_Inf, 'm*-')        # Genera el gráfico
    Representacion_Nivel('Diferencia estandarizada',
    ('DnT_F1_Inf', 'DnT_F2_Inf', 'DnT_Inf'))


    #REPRESENTACIÓN DE R'
    figure("R' de la  Habitación Inferior")
    plot(FR, R_F1_Inf, 'co-', R_F2_Inf, 'bo-', R_Inf, 'm*-')               # Genera el gráfico
    Representacion_Nivel('Índice de reducción sonora aparente',
    ('R_F1_Inf', 'R_F2_Inf', "R'_Inf"))

    tight_layout()                                             # Ajusta la leyenda
    show()