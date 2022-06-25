# TFG Raquel Galán Montes
## Cálculo del Tiempo de Reverberación UNE-EN ISO 16283-1

from openpyxl import load_workbook      # Para leer ficheros .xlsx
import math                             # Para usar funciones matemáticas
from pylab import *                     # Para crear gráficas

# VARIABLES GLOBALES
SHEET = 'TR'        # Hoja del archivo XLSX
FILE_TR = 'RT.xlsx'     # Variable para un segundo fichero (TR)

# ARRAY con constantes
## Rango de frecuencias de interes:
arrayFR = [50, 63, 80, 100, 125, 160, 200, 250, 315, 400, 500, 630, 800, 1000, 1250, 1600, 2000, 2500, 3150, 4000, 5000]

## RANGO DE FRECUENCIAS
FR = ['50', '63', '80', '100', '125', '160', '200', '250', '315', '400', '500', '630', '800', '1000', '1250', '1600', '2000','2500', '3150', '4000', '5000']


# Arrays vacíos
TR = []           # Array para los resultados del tiempo de reverberación de la habitación
TR_F1 = []        # Array para los resultados del tiempo de reverberación en la Posición 1 de la Fuente
TR_F2 = []        # Array para los resultados del tiempo de reverberación en la Posición 2 de la Fuente

# Cálculo del tiempo de reverberación TR60 para cada banda de frecuencia:
def Calcular_TR(A, B, C, D, myArray):
    wb = load_workbook(FILE_TR)        # Se carga en wb el fichero
    sheet = wb[SHEET]                  # Se carga la hoja del fichero de donde obtenemos los datos

    for value in sheet.iter_rows(min_row =A, max_row =B, min_col = C, max_col =D,
                                 values_only=True):
        RT = 0
        i = 0
        n = 0           #NOTA: Se añade un contador de valores por si hubiera algún resultado nulo y 
                        # se tuviera distinto número de promedios.
        for cell in value:
            if value != '*':            # Si los valores no son nulos:
                RT = RT + value[i]      # Se calcula el promedio de los valores del TR por cada posición
                n = n + 1               # Se añade 1 a los valores 'no nulos'
            i = i + 1
        RT = RT/n                       # TR promedio para la frecuencia x
        myArray.append(round(RT,2))     # Array ordenado de datos

# Se calcula el TR promedio de la habitación, donde A: TR_F1 y B: TR_F2
def TR_Habitación(A, B, myArray):
    i = 0   
    for num in A:
        TR = (A[i] + B[i])/2
        i = i + 1 
        myArray.append(round(TR,1))          # Array ordenado de datos

# Se preparan los resultados para imprimirlos por pantalla
def Imprimir(myArray, unidades):
    i = 0                                             # Iterador para los arrays
    for num in arrayFR:
        print(arrayFR[i], 'Hz - ', myArray[i], 's')  # Se imprime por la terminal el elemento i del array
        i = i + 1                                     # Siguiente elemento

# Datos mostrados en una gráfica
def representacion(A):
    title(A)
    xlabel('Frecuencia [Hz]')
    ylabel('Tiempo[s]')
    ylim(0, 1.5)
    legend(('TR60'),
    loc='upper right', facecolor="w", prop = {'size': 15})
    grid()


# Se imprimen todos los datos
if __name__ == "__main__":

    print()
    print('POSICIÓN FUENTE 1')
    print('------------------------')
    print('Frecuencia | TR')
    print('------------------------')
    Calcular_TR(6, 26, 12, 15, TR_F1) # Imprime los valores desde de las filas 6-26 y de las columnas L-O
    Imprimir(TR_F1, 's')

    print()
    print('POSICIÓN FUENTE 2')
    print('------------------------')
    print('Frecuencia | TR')
    print('------------------------')
    Calcular_TR(6, 26, 19, 22, TR_F2) # Imprime los valores desde de las filas 6-26 y de las columnas S-V
    Imprimir(TR_F2, 's')


    print()
    print('HABITACIÓN')
    print('------------------------')
    print('Frecuencia | TR')
    print('------------------------')
    TR_Habitación(TR_F1, TR_F2, TR)
    Imprimir(TR, 's')


    #REPRESENTACIÓN DEL TR
    figure('TR60')
    plot(FR, TR, 'ro-')        # Genera el gráfico
    representacion('Tiempo de reverberación en la sala receptora')

    show()